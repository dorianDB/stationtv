"""
Station TV - Générateur de rapport Excel pour les benchmarks
Convertit les résultats CSV en un fichier Excel formaté avec graphiques.

Usage:
    python scripts/GenerateExcelReport.py --input output/benchmark_results.csv
    
Nécessite: pip install openpyxl pandas
"""

import sys
import argparse
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("❌ Modules requis non installés")
    print("Installez avec: pip install openpyxl pandas")
    sys.exit(1)

# Ajouter le répertoire parent au path
sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.logger import setup_logger

# Logger
logger = setup_logger("GenerateExcelReport", level="INFO")


def load_benchmark_data(csv_file: str) -> pd.DataFrame:
    """
    Charge les données du benchmark depuis le CSV.
    
    Args:
        csv_file: Chemin du fichier CSV
    
    Returns:
        DataFrame pandas
    """
    try:
        df = pd.read_csv(csv_file)
        logger.info(f"✓ Chargé {len(df)} lignes depuis {csv_file}")
        return df
    except Exception as e:
        logger.error(f"❌ Erreur lors du chargement: {str(e)}")
        return None


def create_summary_matrix(df: pd.DataFrame) -> pd.DataFrame:
    """
    Crée une matrice résumé (durée × modèle).
    
    Args:
        df: DataFrame avec les résultats bruts
    
    Returns:
        DataFrame avec la matrice pivot
    """
    # Créer le pivot: durées en lignes, modèles en colonnes
    pivot = df.pivot_table(
        values='avg_time_s',
        index='duration_s',
        columns='model',
        aggfunc='mean'
    )
    
    return pivot


def create_excel_report(csv_file: str, output_file: str):
    """
    Crée un rapport Excel formaté avec graphiques.
    
    Args:
        csv_file: Fichier CSV d'entrée
        output_file: Fichier Excel de sortie
    """
    logger.info("=" * 80)
    logger.info("GÉNÉRATION DU RAPPORT EXCEL")
    logger.info("=" * 80)
    
    # Charger les données
    df = load_benchmark_data(csv_file)
    if df is None:
        return
    
    # Créer le workbook
    wb = Workbook()
    
    # === Feuille 1: Données brutes ===
    ws_raw = wb.active
    ws_raw.title = "Données brutes"
    
    logger.info("Création de la feuille 'Données brutes'...")
    
    # Écrire les données
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_raw.append(r)
    
    # Formater l'en-tête
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws_raw[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Ajuster les largeurs de colonnes
    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_raw.column_dimensions[column_letter].width = adjusted_width
    
    # === Feuille 2: Matrice résumé ===
    ws_matrix = wb.create_sheet("Matrice résumé")
    
    logger.info("Création de la feuille 'Matrice résumé'...")
    
    # Créer la matrice
    matrix = create_summary_matrix(df)
    
    # Ajouter une ligne avec 0 au début (comme dans votre image)
    zero_row = pd.DataFrame([[0] * len(matrix.columns)], columns=matrix.columns, index=[0])
    matrix = pd.concat([zero_row, matrix])
    
    # Écrire la matrice
    ws_matrix.append(['Duration (s)'] + list(matrix.columns))
    for idx, row in matrix.iterrows():
        ws_matrix.append([idx] + list(row))
    
    # Formater l'en-tête
    for cell in ws_matrix[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Calculer Th et 1/Th
    ws_matrix.append([])
    ws_matrix.append(['Metrics'])
    
    th_row = ['Th (avg)']
    inverse_th_row = ['1/Th (avg)']
    
    for model in matrix.columns:
        model_data = df[df['model'] == model]
        if not model_data.empty:
            avg_throughput = model_data['throughput'].mean()
            th_row.append(f"{avg_throughput:.3f}")
            inverse_th_row.append(f"{1/avg_throughput:.3f}" if avg_throughput > 0 else "")
        else:
            th_row.append("")
            inverse_th_row.append("")
    
    ws_matrix.append(th_row)
    ws_matrix.append(inverse_th_row)
    
    # Formater les nombres
    for row in ws_matrix.iter_rows(min_row=2, max_row=ws_matrix.max_row - 3):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.column > 1:
                cell.number_format = '0.00'
    
    # === Feuille 3: Graphiques ===
    ws_charts = wb.create_sheet("Graphiques")
    
    logger.info("Création de la feuille 'Graphiques'...")
    
    # Graphique 1: Temps de traitement vs durée audio
    chart1 = LineChart()
    chart1.title = "Temps de traitement vs Durée audio"
    chart1.x_axis.title = "Durée audio (s)"
    chart1.y_axis.title = "Temps de traitement (s)"
    chart1.width = 20
    chart1.height = 12
    
    # Données pour le graphique (référence à la matrice)
    # Lignes: 2 à max_row-3 (sans les métriques)
    # Colonnes: 2 à max_col (sans la première colonne des durées)
    data_rows = len(matrix)
    data_cols = len(matrix.columns)
    
    data = Reference(ws_matrix, min_col=2, min_row=1, max_col=data_cols+1, max_row=data_rows+1)
    cats = Reference(ws_matrix, min_col=1, min_row=2, max_row=data_rows+1)
    
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    
    # Style des lignes
    chart1.style = 10
    
    ws_charts.add_chart(chart1, "A1")
    
    # Graphique 2: Throughput par modèle
    chart2 = BarChart()
    chart2.title = "Throughput moyen par modèle"
    chart2.x_axis.title = "Modèle"
    chart2.y_axis.title = "Throughput (× temps réel)"
    chart2.width = 15
    chart2.height = 10
    
    # Référence aux données de Th
    th_row_num = ws_matrix.max_row - 1
    data2 = Reference(ws_matrix, min_col=2, max_col=data_cols+1, min_row=th_row_num, max_row=th_row_num)
    cats2 = Reference(ws_matrix, min_col=2, max_col=data_cols+1, min_row=1, max_row=1)
    
    chart2.add_data(data2, titles_from_data=False)
    chart2.set_categories(cats2)
    
    chart2.style = 11
    
    ws_charts.add_chart(chart2, "A25")
    
    # === Feuille 4: Statistiques ===
    ws_stats = wb.create_sheet("Statistiques")
    
    logger.info("Création de la feuille 'Statistiques'...")
    
    ws_stats.append(["Statistiques du Benchmark"])
    ws_stats.append([])
    
    # Statistiques générales
    ws_stats.append(["Métrique", "Valeur"])
    ws_stats.append(["Nombre total de tests", len(df)])
    ws_stats.append(["Nombre de fichiers testés", df['file'].nunique()])
    ws_stats.append(["Nombre de modèles testés", df['model'].nunique()])
    ws_stats.append(["Modèles", ", ".join(df['model'].unique())])
    ws_stats.append([])
    
    # Statistiques par modèle
    ws_stats.append(["Modèle", "Temps moyen (s)", "Temps min (s)", "Temps max (s)", "Throughput moyen"])
    
    for model in sorted(df['model'].unique()):
        model_data = df[df['model'] == model]
        avg_time = model_data['avg_time_s'].mean()
        min_time = model_data['min_time_s'].min()
        max_time = model_data['max_time_s'].max()
        avg_throughput = model_data['throughput'].mean()
        
        ws_stats.append([model, f"{avg_time:.2f}", f"{min_time:.2f}", f"{max_time:.2f}", f"{avg_throughput:.2f}"])
    
    # Formater l'en-tête de la table
    for cell in ws_stats[3]:
        cell.fill = header_fill
        cell.font = header_font
    
    for cell in ws_stats[10]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Sauvegarder le fichier
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb.save(output_file)
    logger.info(f"✅ Rapport Excel créé: {output_file}")


def main():
    """Fonction principale."""
    parser = argparse.ArgumentParser(
        description="Génération de rapport Excel pour les benchmarks Whisper"
    )
    parser.add_argument(
        '--input', '-i',
        default='output/benchmark_results.csv',
        help="Fichier CSV d'entrée (défaut: output/benchmark_results.csv)"
    )
    parser.add_argument(
        '--output', '-o',
        default='output/benchmark_report.xlsx',
        help="Fichier Excel de sortie (défaut: output/benchmark_report.xlsx)"
    )
    
    args = parser.parse_args()
    
    # Vérifier que le fichier d'entrée existe
    if not Path(args.input).exists():
        logger.error(f"❌ Fichier d'entrée introuvable: {args.input}")
        logger.info("Lancez d'abord le benchmark:")
        logger.info("  python scripts/BenchmarkModels.py")
        return
    
    # Créer le rapport
    create_excel_report(args.input, args.output)
    
    logger.info("")
    logger.info("=" * 80)
    logger.info("✅ RAPPORT EXCEL GÉNÉRÉ AVEC SUCCÈS")
    logger.info("=" * 80)
    logger.info(f"📁 Fichier: {args.output}")
    logger.info("")
    logger.info("Le fichier contient 4 feuilles:")
    logger.info("  1. Données brutes - Tous les résultats détaillés")
    logger.info("  2. Matrice résumé - Format compatible avec votre document")
    logger.info("  3. Graphiques - Visualisations automatiques")
    logger.info("  4. Statistiques - Résumé des performances")


if __name__ == "__main__":
    main()
